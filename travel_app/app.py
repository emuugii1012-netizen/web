from flask import Flask, render_template, request, redirect, url_for, send_file, flash
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import re

app = Flask(__name__)
app.secret_key = 'your-secret-key-here-change-in-production'

# Excel файлын нэр
EXCEL_FILE = 'registrations.xlsx'

# Аяллын чиглэл ба огноо
ROUTES = {
    'Датун–Тайланд': '2026-02-15',
    'Сингапур–Индонез': '2026-03-15',
    'Улаанбаатар–Бээжин': '2026-05-30',
    'Бангкок–Паттая': '2026-04-20',
    'Япон–Ёкохама–Фужи уул–Токио': '2026-05-20',
    'Анталя–Памуккале–Истанбул': '2026-06-15'
}

# Аяллын багцууд (хүний тоо ба үнэ)
PACKAGES = {
    'Датун–Тайланд': {
        '1 хүн': 3550000,
        '2 хүн': 5680000
    },
    'Сингапур–Индонез': {
        '1 хүн': 11900000,
        '2 хүн': 19040000
    },
    'Улаанбаатар–Бээжин': {
        '1 хүн': 2690000,
        '2 хүн': 4304000
    },
    'Бангкок–Паттая': {
        '1 хүн': 5490000,
        '2 хүн': 8784000
    },
    'Япон–Ёкохама–Фужи уул–Токио': {
        '1 хүн': 3100000,
        '2 хүн': 4960000
    },
    'Анталя–Памуккале–Истанбул': {
        '1 хүн': 4550000,
        '2 хүн': 7280000
    }
}


def init_excel():
    """Excel файл байхгүй бол үүсгэнэ"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        # Header-ууд үүсгэх
        ws.append([
            'Timestamp',
            'Full Name',
            'Phone',
            'Email',
            'Route',
            'Departure Date',
            'Package',
            'Total Price',
            '50% Deposit',
            'Deposit Confirmed'
        ])
        wb.save(EXCEL_FILE)


def save_to_excel(data):
    """Бүртгэлийг Excel-д хадгална"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([
            data['timestamp'],
            data['full_name'],
            data['phone'],
            data['email'],
            data['route'],
            data['departure_date'],
            data['package'],
            data['total_price'],
            data['deposit'],
            data['deposit_confirmed']
        ])
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        return False


def validate_phone(phone):
    """Утасны дугаар зөвхөн 8 оронтой тоо эсэхийг шалгах"""
    pattern = r'^\d{8}$'
    return re.match(pattern, phone) is not None


@app.route('/')
def index():
    """Үндсэн бүртгэлийн хуудас"""
    return render_template('index.html', routes=ROUTES, packages=PACKAGES)


@app.route('/submit', methods=['POST'])
def submit():
    """Бүртгэл илгээх"""
    try:
        # Form-оос өгөгдөл авах
        full_name = request.form.get('full_name', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()
        route = request.form.get('route', '')
        package = request.form.get('package', '')
        deposit_confirmed = request.form.get('deposit_confirmed', '')

        # Утасны дугаарын валидаци
        if not validate_phone(phone):
            flash('Утасны дугаар зөвхөн 8 оронтой тоо байх ёстой!', 'error')
            return redirect(url_for('index'))

        # Урьдчилгаа төлсөн эсэхийг шалгах
        if deposit_confirmed != 'on':
            flash('Та урьдчилгаа төлсөн эсэхээ баталгаажуулна уу!', 'error')
            return redirect(url_for('index'))

        # Бүх талбарууд бөглөгдсөн эсэхийг шалгах
        if not all([full_name, phone, email, route, package]):
            flash('Бүх талбаруудыг бөглөнө үү!', 'error')
            return redirect(url_for('index'))

        # Үнийг тооцоолох
        total_price = PACKAGES[route][package]
        deposit = total_price / 2
        departure_date = ROUTES[route]

        # Өгөгдөл бэлтгэх
        data = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'full_name': full_name,
            'phone': phone,
            'email': email,
            'route': route,
            'departure_date': departure_date,
            'package': package,
            'total_price': total_price,
            'deposit': deposit,
            'deposit_confirmed': 'Yes'
        }

        # Excel-д хадгалах
        if save_to_excel(data):
            flash('Бүртгэл амжилттай хадгалагдлаа! Баярлалаа.', 'success')
        else:
            flash('Алдаа гарлаа. Дахин оролдоно уу.', 'error')

        return redirect(url_for('index'))

    except Exception as e:
        print(f"Error in submit: {e}")
        flash('Алдаа гарлаа. Дахин оролдоно уу.', 'error')
        return redirect(url_for('index'))


@app.route('/admin')
def admin():
    """Админы хуудас - бүх бүртгэлүүдийг харуулах"""
    try:
        if not os.path.exists(EXCEL_FILE):
            registrations = []
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            registrations = []
            
            # Header-ыг алгасаад бүх мөрийг уншина
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Хоосон мөр биш бол
                    registrations.append({
                        'timestamp': row[0],
                        'full_name': row[1],
                        'phone': row[2],
                        'email': row[3],
                        'route': row[4],
                        'departure_date': row[5],
                        'package': row[6],
                        'total_price': f"{row[7]:,}",
                        'deposit': f"{row[8]:,}",
                        'deposit_confirmed': row[9]
                    })
        
        return render_template('admin.html', registrations=registrations)
    
    except Exception as e:
        print(f"Error in admin: {e}")
        return f"Алдаа гарлаа: {e}"


@app.route('/download')
def download():
    """Excel файл татах"""
    try:
        if os.path.exists(EXCEL_FILE):
            return send_file(
                EXCEL_FILE,
                as_attachment=True,
                download_name=f'registrations_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )
        else:
            flash('Excel файл олдсонгүй!', 'error')
            return redirect(url_for('admin'))
    except Exception as e:
        print(f"Error downloading file: {e}")
        return f"Алдаа гарлаа: {e}"


if __name__ == '__main__':
    # Excel файл үүсгэх
    init_excel()
    # Flask апп ажиллуулах
    app.run(debug=True, host='0.0.0.0', port=5000)
